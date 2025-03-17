import pandas as pd
from pathlib import Path 
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

# Configurações de caminho
BASE_DIR = Path(__file__).parent.parent 
DATA_DIR = BASE_DIR / 'data/raw_data' 
OUTPUT_DIR = BASE_DIR / 'data/processed_data'

# Lista de arquivos de vendas
sales_files = [
    DATA_DIR / 'Meganium_Sales_Data_-_AliExpress.csv',
    DATA_DIR / 'Meganium_Sales_Data_-_Shopee.csv',
    DATA_DIR / 'Meganium_Sales_Data_-_Etsy.csv'
]

# Cria a pasta de saída caso não exista
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def read_delimited_file(path):
    """Lê arquivos delimitados por vírgula ou tabulação."""
    return pd.read_csv(
        path,
        sep='[,|\t]',
        engine='python',
        encoding='utf-8'
    )

def auto_fit_columns(excel_file, sheet_name='Vendas'):
    """Ajusta automaticamente a largura das colunas no Excel."""
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_file)

def generate_reports(df): 
    """Gera relatórios de análise de dados."""
    # Produto mais vendido
    sales_by_product = df.groupby('product_sold', as_index=False)['quantity'].sum()
    sales_by_product = sales_by_product.sort_values('quantity', ascending=False)

    # Vendas por país
    sales_by_country = df.groupby('delivery_country', as_index=False)['quantity'].sum()

    # Ticket médio por moeda
    average_ticket = df.groupby('currency', as_index=False)['total_price'].mean()
    average_ticket['total_price'] = average_ticket['total_price'].round(2)

    return {
        'top_produtos': sales_by_product,
        'vendas_paises': sales_by_country,
        'ticket_medio': average_ticket
    }

def export_analysis(excel_file, analysis):
    """Exporta análises para planilhas separadas no Excel."""
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        analysis['top_produtos'].to_excel(writer, sheet_name='Top Produtos', index=False)
        analysis['vendas_paises'].to_excel(writer, sheet_name='Vendas por País', index=False)
        analysis['ticket_medio'].to_excel(writer, sheet_name='Ticket Médio', index=False)

if __name__ == "__main__":
   
    dfs = [read_delimited_file(file) for file in sales_files]
    combined_files = pd.concat(dfs, ignore_index=True)

    output_file = OUTPUT_DIR / 'Meganium_Sales_Data.xlsx'
    combined_files.to_excel(output_file, index=False, sheet_name='Vendas') 
    auto_fit_columns(output_file)

  
    analysis = generate_reports(combined_files)
    export_analysis(output_file, analysis)

    print(f'''
Arquivo gerado: {output_file}

Conteúdo:
- Planilha "Vendas": Dados brutos consolidados
- Planilha "Top Produtos": Ranking de produtos mais vendidos
- Planilha "Vendas por País": Distribuição geográfica das vendas
- Planilha "Ticket Médio": Valor médio por transação por moeda

Total de registros processados: {len(combined_files):,}
''')